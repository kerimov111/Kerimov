import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
from textwrap import fill
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


dictionary_keys = {'name': 'Название', 'description': 'Описание', 'key_skills': 'Навыки',
                   'experience_id': 'Опыт работы', 'premium': 'Премиум-вакансия',
                   'employer_name': 'Компания', 'salary_from': 'Нижняя граница вилки оклада',
                   'salary_to': 'Верхняя граница вилки оклада', 'salary_gross': 'Оклад указан до вычета налогов',
                   'salary_currency': 'Идентификатор валюты оклада', 'area_name': 'Название региона',
                   'published_at': 'Дата публикации вакансии', 'salary': 'Оклад'}

dictionary_experience_id = {'noExperience': 'Нет опыта', 'between1And3': 'От 1 года до 3 лет',
                            'between3And6': 'От 3 до 6 лет', 'moreThan6': 'Более 6 лет'}

dictionary_salary_currency = {'AZN': 'Манаты', 'BYR': 'Белорусские рубли', 'EUR': 'Евро',
                              'GEL': 'Грузинский лари', 'KGS': 'Киргизский сом',
                              'KZT': 'Тенге', 'RUR': 'Рубли', 'UAH': 'Гривны',
                              'USD': 'Доллары', 'UZS': 'Узбекский сум'}

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

dict_experience_id = {
    'noExperience': 0,
    'between1And3': 1,
    'between3And6': 2,
    'moreThan6': 3,
}

true_false = {
    'False': 'Нет',
    'True': 'Да'
}

list_analytical_dict_year = []
list_analytical_dict_city = []
list_analytical_dict_city_1 = []


class Vacancy:
    def __init__(self, name, salary, area_name, published_at):
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def get_salary_in_rub(self):
        return (float(self.salary_from) + float(self.salary_to)) / 2 * currency_to_rub[self.salary_currency]


class InputConect:
    def print_data(self):
        input_params = InputConect.input_params()
        if input_params is not None:
            file_name, vacancy_name = input_params
            vacancies_objects = DataSet(file_name).vacancies_objects
            InputConect.print_analytical_data(vacancies_objects, vacancy_name)

    @staticmethod
    def input_params():
        file_name = input('Введите название файла: ')
        vacancy_name = input('Введите название профессии: ')

        return file_name, vacancy_name

    @staticmethod
    def get_years_salary_dict(dictionary):
        for key, value in dictionary.items():
            if len(value) == 0:
                dictionary[key] = 0
            else:
                dictionary[key] = int(sum(value) / len(value))
        return dictionary

    @staticmethod
    def print_analytical_data(vacancies_objects, vacancy_name):
        vacancies_dict = vacancies_objects
        years = set()
        for vacancy in vacancies_dict:
            years.add(int(vacancy.published_at[:4]))
        years = list(range(min(years), max(years) + 1))

        years_salary_dictionary = {year: [] for year in years}
        years_salary_vacancy_dict = {year: [] for year in years}
        years_count_dictionary = {year: 0 for year in years}
        years_count_vacancy_dict = {year: 0 for year in years}

        area_dict = {}

        for vacancy in vacancies_dict:
            year = int(vacancy.published_at[:4])
            years_salary_dictionary[year].append(vacancy.salary.get_salary_in_rub())
            years_count_dictionary[year] += 1
            if vacancy_name in vacancy.name:
                years_salary_vacancy_dict[year].append(vacancy.salary.get_salary_in_rub())
                years_count_vacancy_dict[year] += 1
            if vacancy.area_name in area_dict:
                area_dict[vacancy.area_name].append(vacancy.salary.get_salary_in_rub())
            else:
                area_dict[vacancy.area_name] = [vacancy.salary.get_salary_in_rub()]

        years_salary_dictionary = InputConect.get_years_salary_dict(years_salary_dictionary)
        years_salary_vacancy_dict = InputConect.get_years_salary_dict(years_salary_vacancy_dict)

        area_list = area_dict.items()
        area_list = [area for area in area_list if len(area[1]) >= len(vacancies_dict) // 100]
        area_salary_dict = sorted(area_list, key=lambda area: sum(area[1]) / len(area[1]), reverse=True)
        area_count_dict = sorted(area_list, key=lambda item: len(item[1]) / len(vacancies_dict), reverse=True)
        area_salary_dict = {item[0]: int(sum(item[1]) / len(item[1]))
                            for item in area_salary_dict[0: min(len(area_salary_dict), 10)]}
        area_count_dict = {item[0]: round(len(item[1]) / len(vacancies_dict), 4)
                           for item in area_count_dict[0: min(len(area_count_dict), 10)]}

        print(f'Динамика уровня зарплат по годам: {years_salary_dictionary}')
        print(f'Динамика количества вакансий по годам: {years_count_dictionary}')
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {years_salary_vacancy_dict}')
        print(f'Динамика количества вакансий по годам для выбранной профессии: {years_count_vacancy_dict}')
        print(f'Уровень зарплат по городам (в порядке убывания): {area_salary_dict}')
        print(f'Доля вакансий по городам (в порядке убывания): {area_count_dict}')

        list_analytical_dict_year.append(years_salary_dictionary)
        list_analytical_dict_year.append(years_salary_vacancy_dict)
        list_analytical_dict_year.append(years_count_dictionary)
        list_analytical_dict_year.append(years_count_vacancy_dict)
        list_analytical_dict_city.append(area_salary_dict)
        list_analytical_dict_city_1.append(area_count_dict)


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet.csv_reader(file_name)

    @staticmethod
    def delete_tags(value):
        temp_value = ''
        while value.find('<') != - 1:
            temp_value += value[:value.find('<')]
            current_index = value.find('>') + 1
            value = value[current_index:]
        else:
            return temp_value + value

    @staticmethod
    def csv_reader(file_name):
        with open(file_name, newline='', encoding='utf-8-sig') as file:
            vacancies_csv = csv.reader(file)
            vacancy_data = [row for row in vacancies_csv]
            vacancy_keys = []
            try:
                vacancy_keys = vacancy_data.pop(0)
            except:
                print('Пустой файл')
                exit()
            vacancy_dictionary = []
            filtered_vacancy_data = [vacancy for vacancy in vacancy_data
                                     if len(vacancy) == len(vacancy_keys) and '' not in vacancy]
            for row in filtered_vacancy_data:
                dic = {}
                for i in range(len(row)):
                    elem = DataSet.delete_tags(row[i])
                    if elem.find("\n") != -1:
                        elem = elem.split('\n')
                        elem = [' '.join(x.split()) for x in elem]
                    else:
                        elem = ' '.join(elem.split())
                    dic[vacancy_keys[i]] = elem
                vacancy_dictionary.append(
                    Vacancy(dic['name'], Salary(dic['salary_from'], dic['salary_to'],
                            dic['salary_currency']), dic['area_name'], dic['published_at']))
            return vacancy_dictionary


class Report:
    def __init__(self, sheet_title_year, sheet_title_city, color_border, style_border, bold_text):
        self.sheet_title_year = sheet_title_year
        self.sheet_title_city = sheet_title_city
        self.color_border = color_border
        self.style_border = style_border
        self.bold_text = bold_text

    @staticmethod
    def generate_pie_chart(dictionary, title, ax):
        labels = ['Другие']
        for city in dictionary.keys():
            labels.append(city)
        sizes = [1 - sum(dictionary.values())]
        for size in dictionary.values():
            sizes.append(size)
        textprops = {"fontsize": 6}

        ax.set_title(title)
        ax.pie(sizes, labels=labels, textprops=textprops)
        ax.axis('equal')

    @staticmethod
    def generate_horizontal_bar_chart(dictionary, title, ax):
        plt.rcdefaults()

        cities = []
        for city in dictionary.keys():
            if ' ' in city:
                cities.append(city.replace(' ', '\n'))
            elif '-' in city:
                cities.append(city.replace('-', '-\n'))
            else:
                cities.append(city)
        y_pos = np.arange(len(cities))
        performance = dictionary.values()

        ax.barh(y_pos, performance)
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.set_title(title)
        ax.tick_params(axis='x', labelsize=8)
        ax.tick_params(axis='y', labelsize=6)
        ax.grid(visible=True, axis='x')

    @staticmethod
    def generate_group_bar_chart(first_dict, second_dict, first_label, second_label, title, vacancy_name, ax):
        labels = first_dict.keys()
        allVacancy = first_dict.values()
        oneVacancy = second_dict.values()
        x = np.arange(len(labels))

        width = 0.35
        ax.bar(x - width / 2, allVacancy, width, label=first_label)
        ax.bar(x + width / 2, oneVacancy, width, label=fill(f'{second_label} {vacancy_name.lower()}', 20))

        ax.set_title(title)
        ax.set_xticks(x, labels, rotation=90)
        ax.legend(fontsize=8, loc='upper left')
        ax.tick_params(axis='both', labelsize=8)
        ax.grid(visible=True, axis='y')

    @staticmethod
    def generate_image(vacancy_name):
        fig, ax = plt.subplots(2, 2)
        Report.generate_group_bar_chart(list_analytical_dict_year[0], list_analytical_dict_year[1], 'средняя з/п',
                                        'з/п', 'Уровень зарплат по годам', vacancy_name, ax[0][0])
        Report.generate_group_bar_chart(list_analytical_dict_year[2], list_analytical_dict_year[3],
                                        'Количество вакансий', 'Количество вакансий', 'Количество вакансий по годам',
                                        vacancy_name, ax[0][1])
        Report.generate_horizontal_bar_chart(list_analytical_dict_city[0], 'Уровень зарплат по городам', ax[1][0])
        Report.generate_pie_chart(list_analytical_dict_city_1[0], 'Доля вакансий по городам', ax[1][1])
        plt.tight_layout()
        plt.savefig('graph.png')
        plt.show()



    @staticmethod
    def append_values(list_dict, work_sheet, style_border, num_column=1):
        for i, elem in enumerate(list_dict[0].keys()):
            work_sheet.cell(row=i + 2, column=num_column).value = elem
            work_sheet.cell(row=i + 2, column=num_column).border = style_border
        column_num = num_column + 1
        for dic in list_dict:
            for i, elem in enumerate(dic.values()):
                if type(elem) == float:
                    work_sheet.cell(row=i + 2, column=column_num).number_format = '0.00%'
                work_sheet.cell(row=i + 2, column=column_num).value = elem
                work_sheet.cell(row=i + 2, column=column_num).border = style_border
            column_num += 1

    @staticmethod
    def get_width_column(sheet, font_size=11):
        cols_dict = {}
        for row in sheet.rows:
            for cell in row:
                letter = cell.column_letter
                if cell.value:
                    len_cell = len(str(cell.value))
                    len_cell_dict = 0
                    if letter in cols_dict:
                        len_cell_dict = cols_dict[letter]
                    if len_cell > len_cell_dict:
                        cols_dict[letter] = len_cell
                        new_width_col = len_cell * font_size ** (font_size * 0.009)
                        sheet.column_dimensions[cell.column_letter].width = new_width_col

    def generate_excel(self, vacancy_name):
        field_statistic_year = ['Год', 'Средняя зарплата', f'Средняя зарплата - {vacancy_name}',
                                'Количество вакансий', f'Количество вакансий - {vacancy_name}']
        field_statistic_city = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        style_border = Side(border_style=self.style_border, color=self.color_border)
        bold_heading = Font(bold=self.bold_text)
        border = Border(top=style_border, bottom=style_border, left=style_border, right=style_border)
        work_book = Workbook()
        work_sheet_year = work_book.active
        work_sheet_year.title = self.sheet_title_year
        work_sheet_city = work_book.create_sheet(self.sheet_title_city)
        work_sheet_year.append(field_statistic_year)
        work_sheet_city.append(field_statistic_city)
        Report.append_values(list_analytical_dict_year, work_sheet_year, border)
        Report.append_values(list_analytical_dict_city, work_sheet_city, border)
        Report.append_values(list_analytical_dict_city_1, work_sheet_city, border, 3)
        for sheet in work_book:
            for cell in sheet[1]:
                cell.font = bold_heading
                cell.border = border
        work_sheet_city.insert_cols(3, 1)

        Report.get_width_column(work_sheet_year)
        Report.get_width_column(work_sheet_city)

        work_book.save('report.xlsx')

    @staticmethod
    def generate_pdf(vacancy_name):
        field_table_city_1 = ['Город', 'Уровень зарплат']
        field_table_city_2 = ['Город', 'Доля вакансий']

        field_statistic_year = ['Год', 'Средняя зарплата', f'Средняя зарплата - {vacancy_name}',
                                'Количество вакансий', f'Количество вакансий - {vacancy_name}']

        years = list_analytical_dict_year[0].keys()

        table_name_1 = 'Статистика по годам'
        table_name_2 = 'Статистика по городам'

        image_url = 'C:\\Users\\elena\\PycharmProjects\\LatyntsevaElena\\graph.png'

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("index.html")

        dict_city_part = {}

        for key, value in list_analytical_dict_city_1[0].items():
            value = f'{round(value * 100, 2)}%'
            dict_city_part[key] = value

        pdf_template = template.render({'vacancy_name': vacancy_name, 'image_url': image_url,
                                        'field_statistic_year': field_statistic_year, 'years': years,
                                        'list_analytical_dict_year': list_analytical_dict_year,
                                        'table_name_1': table_name_1, 'table_name_2': table_name_2,
                                        'field_table_city_1': field_table_city_1,
                                        'field_table_city_2': field_table_city_2,
                                        'list_analytical_dict_city': list_analytical_dict_city[0],
                                        'dict_city_part': dict_city_part})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


def main():
    a = InputConect()
    a.print_data()
